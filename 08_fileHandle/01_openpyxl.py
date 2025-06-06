import openpyxl
book = openpyxl.load_workbook("C:/Users/suyas/Desktop/selenium/08_fileHandle/new_test.xlsx") #path

sheet = book.active #access sheet

Dict = {}

cell = sheet.cell(row=6, column=1) #access cell
print(cell.value)  

sheet.cell(row=3, column=5).value = "Suyash"  #update value
print(sheet.cell(row=1, column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A7'].value)

print("--------------------------------")

# #print all values
# for i in range(1, sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         print(sheet.cell(row=i, column=j).value)

# #print values of tc2
# for i in range(1, sheet.max_row+1):
#     if sheet.cell(row=i, column=1).value == "tc2":
#         for j in range(1, sheet.max_column+1):
#             print(sheet.cell(row=i, column=j).value)

#print values of tc2 in dictionary
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "tc2":
        for j in range(1, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)

